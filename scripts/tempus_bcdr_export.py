#!/usr/bin/env python3
"""
BCDR Export - Tempus Disaster Recovery
======================================
Standalone script to extract all data from the Tempus database
and generate an Excel file (.xlsx) per user with:
  - "Time Entries" tab: clock-in/clock-out records
  - "Holidays and Absences" tab: holiday and leave requests

Usage:
    python scripts/bcdr_export.py
    python scripts/bcdr_export.py --output /destination/path
    python scripts/bcdr_export.py --host localhost --port 5432 --db fichador_db --user fichador_user --password fichador_pass

Requirements:
    pip install psycopg2-binary openpyxl
"""

import argparse
import os
import sys
from datetime import datetime, date, time, timedelta
from pathlib import Path

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("ERROR: psycopg2 not installed. Run: pip install psycopg2-binary")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ─── Excel Styles ────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_COLORS = {
    "aprobada": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "pendiente": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "rechazada": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def calcular_horas(hora_entrada, hora_salida, pausa_min):
    """Calculate worked hours, matching the Fichaje.horas_trabajadas() model method."""
    if not hora_entrada or not hora_salida:
        return None

    entrada_min = hora_entrada.hour * 60 + hora_entrada.minute
    salida_min = hora_salida.hour * 60 + hora_salida.minute

    # Night shift (clock-out on the next day)
    if salida_min < entrada_min:
        diff = (1440 - entrada_min) + salida_min
    else:
        diff = salida_min - entrada_min

    diff -= (pausa_min or 0)
    horas = max(diff / 60.0, 0)
    return round(horas, 2)


def formato_hora(t):
    """Format a time object to HH:MM."""
    if t is None:
        return ""
    if isinstance(t, time):
        return t.strftime("%H:%M")
    return str(t)


def formato_fecha(d):
    """Format a date/datetime to DD/MM/YYYY."""
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def sanitize_filename(name):
    """Generate a safe filename."""
    return "".join(c if c.isalnum() or c in " _-" else "_" for c in name).strip()


def apply_header_style(ws, row, num_cols):
    """Apply header style to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_data_style(cell, center=True):
    """Apply data style to a cell."""
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGNMENT if center else DATA_ALIGNMENT_LEFT
    cell.border = THIN_BORDER


def auto_width(ws):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 10), 40)


# ─── DB Connection ───────────────────────────────────────────────────────────

def connect_db(args):
    """Connect to PostgreSQL with the given parameters."""
    params = {
        "host": args.host,
        "port": args.port,
        "dbname": args.db,
        "user": args.user,
        "password": args.password,
    }
    print(f"Connecting to PostgreSQL {params['host']}:{params['port']}/{params['dbname']}...")
    try:
        conn = psycopg2.connect(**params, cursor_factory=psycopg2.extras.DictCursor)
        print("  Connection established.")
        return conn
    except psycopg2.Error as e:
        print(f"Connection ERROR: {e}")
        sys.exit(1)


# ─── Queries ─────────────────────────────────────────────────────────────────

def get_usuarios(conn):
    """Fetch all active users."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT id, nombre, email, rol, dias_vacaciones, fecha_alta, activo
            FROM usuarios
            ORDER BY nombre
        """)
        return cur.fetchall()


def get_fichajes(conn, usuario_id):
    """Fetch current time entries for a user, sorted by date."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT f.fecha, f.hora_entrada, f.hora_salida, f.pausa,
                   f.tipo_accion, f.motivo_rectificacion,
                   e.nombre as editor_nombre
            FROM fichajes f
            LEFT JOIN usuarios e ON f.editor_id = e.id
            WHERE f.usuario_id = %s
              AND f.es_actual = true
              AND f.tipo_accion != 'eliminacion'
            ORDER BY f.fecha ASC, f.hora_entrada ASC
        """, (usuario_id,))
        return cur.fetchall()


def get_vacaciones(conn, usuario_id):
    """Fetch current holiday requests for a user."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT sv.fecha_inicio, sv.fecha_fin, sv.dias_solicitados,
                   sv.motivo, sv.estado, sv.fecha_solicitud, sv.fecha_respuesta,
                   sv.comentarios, sv.tipo_accion,
                   a.nombre as aprobador_nombre
            FROM solicitudes_vacaciones sv
            LEFT JOIN usuarios a ON sv.aprobador_id = a.id
            WHERE sv.usuario_id = %s
              AND sv.es_actual = true
              AND sv.tipo_accion != 'eliminacion'
            ORDER BY sv.fecha_inicio ASC
        """, (usuario_id,))
        return cur.fetchall()


def get_bajas(conn, usuario_id):
    """Fetch current leave/absence requests for a user."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT sb.fecha_inicio, sb.fecha_fin, sb.dias_solicitados,
                   sb.motivo, sb.estado, sb.fecha_solicitud, sb.fecha_respuesta,
                   sb.comentarios,
                   ta.nombre as tipo_ausencia,
                   a.nombre as aprobador_nombre
            FROM solicitudes_bajas sb
            LEFT JOIN tipos_ausencia ta ON sb.tipo_ausencia_id = ta.id
            LEFT JOIN usuarios a ON sb.aprobador_id = a.id
            WHERE sb.usuario_id = %s
              AND sb.es_actual = true
            ORDER BY sb.fecha_inicio ASC
        """, (usuario_id,))
        return cur.fetchall()


def get_saldos_vacaciones(conn, usuario_id):
    """Fetch holiday balances by year."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT anio, dias_totales, dias_disfrutados, dias_carryover
            FROM saldos_vacaciones
            WHERE usuario_id = %s
            ORDER BY anio ASC
        """, (usuario_id,))
        return cur.fetchall()


# ─── Excel Generation ────────────────────────────────────────────────────────

def crear_hoja_fichajes(wb, fichajes, usuario_nombre):
    """Create the time entries tab."""
    ws = wb.active
    ws.title = "Fichajes"

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Registro de Fichajes — {usuario_nombre}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Export info
    ws.merge_cells("A2:G2")
    info_cell = ws["A2"]
    info_cell.value = f"Exported on {datetime.now().strftime('%d/%m/%Y at %H:%M')}"
    info_cell.font = Font(name="Calibri", italic=True, size=9, color="666666")
    info_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Fecha", "Entrada", "Salida", "Pausa (min)", "Horas Trabajadas", "Tipo", "Observaciones"]
    row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    apply_header_style(ws, row, len(headers))
    ws.row_dimensions[row].height = 25

    # Data
    total_horas = 0
    total_dias = 0

    if not fichajes:
        ws.merge_cells(f"A5:G5")
        empty_cell = ws["A5"]
        empty_cell.value = "No time entries recorded"
        empty_cell.font = Font(name="Calibri", italic=True, color="999999", size=10)
        empty_cell.alignment = Alignment(horizontal="center")
    else:
        for i, f in enumerate(fichajes):
            r = row + 1 + i
            horas = calcular_horas(f["hora_entrada"], f["hora_salida"], f["pausa"])

            ws.cell(row=r, column=1, value=formato_fecha(f["fecha"]))
            ws.cell(row=r, column=2, value=formato_hora(f["hora_entrada"]))
            ws.cell(row=r, column=3, value=formato_hora(f["hora_salida"]))
            ws.cell(row=r, column=4, value=f["pausa"] or 0)
            ws.cell(row=r, column=5, value=f"{horas:.2f}h" if horas is not None else "In progress")
            tipo = "Rectification" if f["tipo_accion"] == "rectificacion" else "Original"
            ws.cell(row=r, column=6, value=tipo)
            obs = f["motivo_rectificacion"] or ""
            if f["editor_nombre"]:
                obs = f"Edited by {f['editor_nombre']}" + (f" — {obs}" if obs else "")
            ws.cell(row=r, column=7, value=obs)

            for col in range(1, len(headers) + 1):
                center = col != 7
                apply_data_style(ws.cell(row=r, column=col), center=center)

            if horas is not None:
                total_horas += horas
            total_dias += 1

        # Summary row
        r_sum = row + 1 + len(fichajes) + 1
        ws.merge_cells(f"A{r_sum}:C{r_sum}")
        summary = ws.cell(row=r_sum, column=1, value="TOTAL")
        summary.font = SUBHEADER_FONT
        summary.fill = SUBHEADER_FILL
        summary.alignment = Alignment(horizontal="right", vertical="center")
        for col in range(1, 4):
            ws.cell(row=r_sum, column=col).fill = SUBHEADER_FILL
            ws.cell(row=r_sum, column=col).border = THIN_BORDER

        ws.cell(row=r_sum, column=4, value=f"{total_dias} days")
        ws.cell(row=r_sum, column=4).font = SUBHEADER_FONT
        ws.cell(row=r_sum, column=4).fill = SUBHEADER_FILL
        ws.cell(row=r_sum, column=4).border = THIN_BORDER
        ws.cell(row=r_sum, column=4).alignment = DATA_ALIGNMENT

        ws.cell(row=r_sum, column=5, value=f"{total_horas:.2f}h")
        ws.cell(row=r_sum, column=5).font = SUBHEADER_FONT
        ws.cell(row=r_sum, column=5).fill = SUBHEADER_FILL
        ws.cell(row=r_sum, column=5).border = THIN_BORDER
        ws.cell(row=r_sum, column=5).alignment = DATA_ALIGNMENT

        for col in range(6, len(headers) + 1):
            ws.cell(row=r_sum, column=col).fill = SUBHEADER_FILL
            ws.cell(row=r_sum, column=col).border = THIN_BORDER

    auto_width(ws)
    ws.sheet_properties.tabColor = "2F5496"
    return ws


def crear_hoja_ausencias(wb, vacaciones, bajas, saldos, usuario_nombre):
    """Create the holidays and absences tab."""
    ws = wb.create_sheet("Vacaciones y Ausencias")

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Vacaciones y Ausencias — {usuario_nombre}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3

    # ── Section: Holiday Balance ──
    if saldos:
        ws.merge_cells(f"A{row}:F{row}")
        sec = ws.cell(row=row, column=1, value="SALDO DE VACACIONES POR AÑO")
        sec.font = Font(name="Calibri", bold=True, size=11, color="2F5496")
        row += 1

        saldo_headers = ["Año", "Días Totales", "Días Disfrutados", "Días Restantes", "Días Carryover"]
        for col, h in enumerate(saldo_headers, 1):
            ws.cell(row=row, column=col, value=h)
        apply_header_style(ws, row, len(saldo_headers))
        row += 1

        for s in saldos:
            restantes = (s["dias_totales"] or 0) - (s["dias_disfrutados"] or 0)
            ws.cell(row=row, column=1, value=s["anio"])
            ws.cell(row=row, column=2, value=s["dias_totales"])
            ws.cell(row=row, column=3, value=s["dias_disfrutados"])
            ws.cell(row=row, column=4, value=restantes)
            ws.cell(row=row, column=5, value=s["dias_carryover"] or 0)
            for col in range(1, len(saldo_headers) + 1):
                apply_data_style(ws.cell(row=row, column=col))
            row += 1

        row += 1

    # ── Section: Holidays ──
    ws.merge_cells(f"A{row}:I{row}")
    sec = ws.cell(row=row, column=1, value="SOLICITUDES DE VACACIONES")
    sec.font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    row += 1

    vac_headers = ["Fecha Inicio", "Fecha Fin", "Días", "Estado", "Motivo",
                   "Fecha Solicitud", "Aprobador", "Fecha Respuesta", "Comentarios"]
    for col, h in enumerate(vac_headers, 1):
        ws.cell(row=row, column=col, value=h)
    apply_header_style(ws, row, len(vac_headers))
    row += 1

    if not vacaciones:
        ws.merge_cells(f"A{row}:I{row}")
        empty_cell = ws.cell(row=row, column=1, value="No holiday requests found")
        empty_cell.font = Font(name="Calibri", italic=True, color="999999", size=10)
        empty_cell.alignment = Alignment(horizontal="center")
        row += 1
    else:
        for v in vacaciones:
            ws.cell(row=row, column=1, value=formato_fecha(v["fecha_inicio"]))
            ws.cell(row=row, column=2, value=formato_fecha(v["fecha_fin"]))
            ws.cell(row=row, column=3, value=v["dias_solicitados"])
            ws.cell(row=row, column=4, value=(v["estado"] or "").capitalize())
            ws.cell(row=row, column=5, value=v["motivo"] or "")
            ws.cell(row=row, column=6, value=formato_fecha(v["fecha_solicitud"]))
            ws.cell(row=row, column=7, value=v["aprobador_nombre"] or "")
            ws.cell(row=row, column=8, value=formato_fecha(v["fecha_respuesta"]))
            ws.cell(row=row, column=9, value=v["comentarios"] or "")

            for col in range(1, len(vac_headers) + 1):
                center = col not in (5, 9)
                apply_data_style(ws.cell(row=row, column=col), center=center)

            # Color by status
            estado = (v["estado"] or "").lower()
            if estado in STATUS_COLORS:
                ws.cell(row=row, column=4).fill = STATUS_COLORS[estado]

            row += 1

    row += 1

    # ── Section: Absences / Leaves ──
    ws.merge_cells(f"A{row}:I{row}")
    sec = ws.cell(row=row, column=1, value="SOLICITUDES DE AUSENCIAS / BAJAS")
    sec.font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    row += 1

    baja_headers = ["Tipo Ausencia", "Fecha Inicio", "Fecha Fin", "Días", "Estado",
                    "Motivo", "Aprobador", "Fecha Respuesta", "Comentarios"]
    for col, h in enumerate(baja_headers, 1):
        ws.cell(row=row, column=col, value=h)
    apply_header_style(ws, row, len(baja_headers))
    row += 1

    if not bajas:
        ws.merge_cells(f"A{row}:I{row}")
        empty_cell = ws.cell(row=row, column=1, value="No absence requests found")
        empty_cell.font = Font(name="Calibri", italic=True, color="999999", size=10)
        empty_cell.alignment = Alignment(horizontal="center")
    else:
        for b in bajas:
            ws.cell(row=row, column=1, value=b["tipo_ausencia"] or "No type")
            ws.cell(row=row, column=2, value=formato_fecha(b["fecha_inicio"]))
            ws.cell(row=row, column=3, value=formato_fecha(b["fecha_fin"]))
            ws.cell(row=row, column=4, value=b["dias_solicitados"])
            ws.cell(row=row, column=5, value=(b["estado"] or "").capitalize())
            ws.cell(row=row, column=6, value=b["motivo"] or "")
            ws.cell(row=row, column=7, value=b["aprobador_nombre"] or "")
            ws.cell(row=row, column=8, value=formato_fecha(b["fecha_respuesta"]))
            ws.cell(row=row, column=9, value=b["comentarios"] or "")

            for col in range(1, len(baja_headers) + 1):
                center = col not in (6, 9)
                apply_data_style(ws.cell(row=row, column=col), center=center)

            estado = (b["estado"] or "").lower()
            if estado in STATUS_COLORS:
                ws.cell(row=row, column=5).fill = STATUS_COLORS[estado]

            row += 1

    auto_width(ws)
    ws.sheet_properties.tabColor = "548235"
    return ws


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="BCDR Export - Extract Tempus data to Excel per user",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/bcdr_export.py
  python scripts/bcdr_export.py --host 192.168.1.100 --port 5432
  python scripts/bcdr_export.py --output /backups/tempus
        """,
    )
    parser.add_argument("--host", default=os.getenv("POSTGRES_HOST", "localhost"),
                        help="PostgreSQL host (default: $POSTGRES_HOST or localhost)")
    parser.add_argument("--port", default=os.getenv("POSTGRES_PORT", "5432"),
                        help="PostgreSQL port (default: 5432)")
    parser.add_argument("--db", default=os.getenv("POSTGRES_DB", "fichador_db"),
                        help="Database name (default: $POSTGRES_DB or fichador_db)")
    parser.add_argument("--user", default=os.getenv("POSTGRES_USER", "fichador_user"),
                        help="DB user (default: $POSTGRES_USER or fichador_user)")
    parser.add_argument("--password", default=os.getenv("POSTGRES_PASSWORD", "fichador_pass"),
                        help="DB password (default: $POSTGRES_PASSWORD or fichador_pass)")
    parser.add_argument("--output", "-o", default=None,
                        help="Output directory (default: bcdr_export_YYYYMMDD_HHMM/)")
    parser.add_argument("--all", action="store_true", default=False,
                        help="Include inactive users")

    args = parser.parse_args()

    # Output directory
    if args.output:
        output_dir = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_dir = Path(f"bcdr_export_{timestamp}")

    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {output_dir.resolve()}\n")

    # Connection
    conn = connect_db(args)

    try:
        usuarios = get_usuarios(conn)
        if not args.all:
            usuarios = [u for u in usuarios if u["activo"]]

        print(f"Users found: {len(usuarios)}\n")
        print("=" * 60)

        exported = 0
        for usuario in usuarios:
            uid = usuario["id"]
            nombre = usuario["nombre"]
            estado = "ACTIVE" if usuario["activo"] else "INACTIVE"

            print(f"\n[{estado}] {nombre} ({usuario['email']})")

            # Fetch data
            fichajes = get_fichajes(conn, uid)
            vacaciones = get_vacaciones(conn, uid)
            bajas = get_bajas(conn, uid)
            saldos = get_saldos_vacaciones(conn, uid)

            print(f"  Time entries: {len(fichajes)} | Holidays: {len(vacaciones)} | "
                  f"Absences: {len(bajas)} | Balances: {len(saldos)}")

            # Create Excel
            wb = Workbook()
            crear_hoja_fichajes(wb, fichajes, nombre)
            crear_hoja_ausencias(wb, vacaciones, bajas, saldos, nombre)

            # Save
            safe_name = sanitize_filename(nombre)
            filename = f"{safe_name}.xlsx"
            filepath = output_dir / filename
            wb.save(filepath)
            print(f"  -> {filepath}")
            exported += 1

        print("\n" + "=" * 60)
        print(f"\nExport completed: {exported} files in {output_dir.resolve()}")

    finally:
        conn.close()
        print("Connection closed.")


if __name__ == "__main__":
    main()
